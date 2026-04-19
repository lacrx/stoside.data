"""Write decoder/oceanside_decoder_output.xlsx matching the Strong Towns
Finance Decoder Worksheet's Input-sheet layout, populated with Oceanside data.

Structure matches the worksheet:
  Row 4: column headers with fiscal years
  Rows 6-11: Net Position inputs
  Row 14: Total Revenues
  Row 15: Operating Grants
  Row 16: Capital Grants
  Row 18: Interest
  Row 19: Net Book TCA
  Row 26: Total Cost TCA (blank — not extracted)
  Rows 29-39: 7 metrics (Sustainability / Flexibility / Vulnerability)
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

ROOT = Path(__file__).parent.parent
MODEL = ROOT / "data_model.json"
OUT_XLSX = ROOT / "decoder" / "oceanside_decoder_output.xlsx"


def main():
    model = json.loads(MODEL.read_text())
    metrics_rows = {r["fiscal_year"]: r for r in model.get("decoder_metrics", [])}
    inputs_rows = {r["fiscal_year"]: r for r in model.get("acfr_decoder_inputs", [])}

    years = sorted(metrics_rows.keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Oceanside Decoder"

    # Header
    ws["A1"] = "City of Oceanside — Strong Towns Finance Decoder"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated from ACFR data in data/budget_history/budget_history.sqlite"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A3"] = "Units: USD. Blank cells = not extracted (see DECODER_ANALYSIS.md)."
    ws["A3"].font = Font(italic=True, size=9)

    # Column headers (matching Decoder layout: year labels in columns E..K)
    row = 5
    ws.cell(row=row, column=1, value="Decoder Line").font = Font(bold=True)
    ws.cell(row=row, column=2, value="ACFR Section").font = Font(bold=True)
    for i, fy in enumerate(years):
        cell = ws.cell(row=row, column=3 + i, value=fy)
        cell.font = Font(bold=True)

    SECTION_FILL = PatternFill("solid", fgColor="E0E8F0")
    METRIC_FILL = PatternFill("solid", fgColor="F5E8D0")

    def write_row(r: int, label: str, acfr_ref: str, fn, style="data"):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=acfr_ref)
        if style == "section":
            for c in range(1, 3 + len(years)):
                ws.cell(row=r, column=c).fill = SECTION_FILL
                ws.cell(row=r, column=c).font = Font(bold=True)
            return
        if style == "metric":
            ws.cell(row=r, column=1).fill = METRIC_FILL
            ws.cell(row=r, column=1).font = Font(bold=True)
        for i, fy in enumerate(years):
            v = fn(fy)
            if v is not None:
                cell = ws.cell(row=r, column=3 + i, value=v)
                if style == "metric_ratio":
                    cell.number_format = '0.00"x"'
                elif style == "metric_pct":
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '"$"#,##0'

    def inp(fy, key):
        return inputs_rows.get(fy, {}).get(key)

    def met(fy, key):
        return metrics_rows.get(fy, {}).get(key)

    # Inputs section
    write_row(5, "", "", None, style="section")
    write_row(6, "  Current Assets (derived)", "MD&A Net position", lambda fy: met(fy, "current_assets"))
    write_row(7, "  Capital Assets (derived)", "MD&A Net position", lambda fy: met(fy, "capital_assets"))
    write_row(8, "Total Assets", "MD&A Net position", lambda fy: met(fy, "total_assets"))
    write_row(9, "  Deferred Outflows", "MD&A Net position", lambda fy: met(fy, "deferred_outflows"))
    write_row(10, "  Liabilities", "MD&A Net position", lambda fy: met(fy, "liabilities"))
    write_row(11, "  Deferred Inflows", "MD&A Net position", lambda fy: met(fy, "deferred_inflows"))
    write_row(12, "Total Liabilities", "(Liabilities + Def. Inflows)", lambda fy: met(fy, "total_liabilities"))
    write_row(13, "", "", None, style="section")
    write_row(14, "Total Revenues", "MD&A Changes in NP (derived)", lambda fy: met(fy, "total_revenues"))
    write_row(15, "  Operating Grants & Contributions", "Gov activities", lambda fy: met(fy, "op_grants"))
    write_row(16, "  Capital Grants & Contributions", "Gov activities", lambda fy: met(fy, "cap_grants"))
    write_row(18, "Interest Charges", "Stmt of Activities", lambda fy: met(fy, "interest_charges"))
    write_row(19, "Net Book TCA", "Cap. assets, net of dep.", lambda fy: met(fy, "net_book_tca"))
    write_row(26, "Total Cost of TCA", "Capital Assets Note (not extracted — see doc)", lambda fy: met(fy, "total_cost_tca"))

    # Metrics section
    write_row(28, "Sustainability Indicators", "", None, style="section")
    write_row(29, "  1. Net Financial Position", "Cur Assets − Total Liab",
              lambda fy: met(fy, "metric_1_net_financial_position"), style="metric")
    write_row(30, "  2. Financial Assets-to-Liabilities", "Cur Assets ÷ Total Liab",
              lambda fy: met(fy, "metric_2_financial_assets_to_liab"), style="metric_ratio")
    write_row(31, "  3. Total Assets-to-Liabilities", "(TA + DO) ÷ Total Liab",
              lambda fy: met(fy, "metric_3_total_assets_to_liab"), style="metric_ratio")
    write_row(32, "  4. Net Debt-to-Total Revenues", "−NetFP ÷ Total Rev (if NetFP<0)",
              lambda fy: met(fy, "metric_4_net_debt_to_revenues"), style="metric_pct")
    write_row(34, "Flexibility Indicators", "", None, style="section")
    write_row(35, "  5. Interest-to-Total Revenues", "Interest ÷ Total Rev",
              lambda fy: met(fy, "metric_5_interest_to_revenues"), style="metric_pct")
    write_row(36, "  6. Net Book-to-Cost of TCA", "Net Book ÷ Gross TCA",
              lambda fy: met(fy, "metric_6_net_book_to_cost_tca"), style="metric_pct")
    write_row(38, "Vulnerability Indicator", "", None, style="section")
    write_row(39, "  7. Govt Transfers-to-Total Revenues", "(Op + Cap Grants) ÷ Total Rev",
              lambda fy: met(fy, "metric_7_transfers_to_revenues"), style="metric_pct")

    # Set column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 32
    for i in range(len(years)):
        ws.column_dimensions[chr(ord("C") + i)].width = 14

    OUT_XLSX.parent.mkdir(exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
