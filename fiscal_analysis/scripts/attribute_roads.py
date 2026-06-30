#!/usr/bin/env python3
"""Attribute road maintenance costs to parcels.

Methodology:
  1. Extract PCI segments from Excel (4,311 segments with street name, PCI, area)
  2. Compute annual lifecycle cost per segment (area × cost_per_sqyd ÷ cycle_years)
  3. Match PCI segments to SanGIS street centerlines by name
  4. For LOCAL streets: attribute cost by parcel frontage (buffer + spatial join)
  5. For COLLECTORS/ARTERIALS: attribute cost by ITE trip generation rates
  6. Output: road cost per parcel

Inputs:
  - capital_improvements/pci/Oceanside PCI.xlsx (PCI segments)
  - capital_improvements/pavement_analysis_verified.json (SCS treatment costs)
  - fiscal_analysis/raw_data/streets_oceanside.geojson (street centerlines)
  - fiscal_analysis/parcels_oceanside.geojson (parcel polygons)

Output:
  - fiscal_analysis/road_costs_by_parcel.json
"""

import json
from collections import defaultdict
from pathlib import Path

import geopandas as gpd
import numpy as np
import openpyxl
from shapely.geometry import LineString

HERE = Path(__file__).parent.parent
PROJECT_ROOT = HERE.parent

PCI_FILE = PROJECT_ROOT / "capital_improvements" / "pci" / "Oceanside PCI.xlsx"
COSTS_FILE = PROJECT_ROOT / "capital_improvements" / "pavement_analysis_verified.json"
STREETS_FILE = HERE / "raw_data" / "streets_oceanside.geojson"
PARCELS_FILE = HERE / "parcels_oceanside.geojson"
OUTPUT = HERE / "road_costs_by_parcel.json"
SUMMARY_OUTPUT = HERE / "road_cost_summary.json"

# SCS 2022 treatment cost per sqyd by PCI bracket and road type
# From pavement_analysis_verified.json
COST_PER_SQYD = {
    "urban_major": {
        "slurry": 5.40, "chip": 7.65, "thin_overlay": 18.0,
        "mill_overlay": 31.5, "fdr": 45.0, "reconstruct": 90.0,
    },
    "urban_local": {
        "slurry": 4.50, "chip": 6.30, "thin_overlay": 16.2,
        "mill_overlay": 27.0, "fdr": 40.5, "reconstruct": 81.0,
    },
}

# Treatment by PCI bracket → which treatment type applies
PCI_TREATMENT = [
    (80, 100, "slurry", 8),       # Excellent: preventive, 8-yr cycle
    (70, 79, "chip", 6),          # Good: chip seal, 6-yr cycle
    (55, 69, "thin_overlay", 8),  # Fair: thin overlay, 8-yr cycle
    (40, 54, "mill_overlay", 10), # At Risk: mill & overlay, 10-yr cycle
    (25, 39, "fdr", 12),          # Poor: full depth reclamation, 12-yr cycle
    (0, 24, "reconstruct", 20),   # Failed: full reconstruction, 20-yr cycle
]

# ITE Trip Generation rates (10th Ed)
ITE_TRIPS = {
    "SFR": 9.44,             # per dwelling unit
    "MFR": 6.74,             # per dwelling unit
    "Commercial": 42.70,     # per 1000 sqft
    "Industrial": 3.89,      # per 1000 sqft
    "Agricultural": 1.0,     # minimal
    "Institutional": 10.0,   # per 1000 sqft (estimate)
    "Government": 10.0,
    "Vacant": 0.0,
    "Open Space": 0.0,
    "Common Area": 0.0,
    "Exempt": 0.0,
    "Other": 1.0,
    "Unknown": 1.0,
    "Recreation": 2.0,
    "Mixed Use": 7.0,        # weighted avg
}

# Map PCI functional class to cost table
FC_TO_COST_TABLE = {
    "Local": "urban_local",
    "Major Arterial": "urban_major",
    "Minor Arterial": "urban_major",
    "Major Collector": "urban_major",
    "Minor Collector": "urban_local",
    "Rural Collector": "urban_local",
    "Rural Local": "urban_local",
}

# Map PCI functional class to local vs arterial
FC_IS_LOCAL = {
    "Local": True,
    "Major Arterial": False,
    "Minor Arterial": False,
    "Major Collector": False,
    "Minor Collector": True,
    "Rural Collector": True,
    "Rural Local": True,
}

# Map SanGIS FUNCLASS codes to broad categories
SANGIS_FC = {
    "L": "Local",
    "C": "Collector",
    "A": "Arterial",
    "2": "Arterial",   # Principal Arterial
    "4": "Arterial",   # Minor Arterial
    "6": "Freeway",
    "7": "Freeway",
    "5": "Freeway",
    "1": "Freeway",
    "F": "Freeway",
    "D": "Local",
    "R": "Ramp",
    "P": "Private",
    "S": "Local",
    "U": "Local",
    "E": "Expressway",
    "T": "Trail",
    "8": "Other",
}


def load_pci_segments() -> list[dict]:
    """Load PCI segments from Excel."""
    wb = openpyxl.load_workbook(PCI_FILE, read_only=True, data_only=True)
    ws = wb["Raw PCI"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    segments = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        name = d.get("Street Name")
        if not name:
            continue
        pci = d.get("Pavement Condition Index (PCI)")
        area = d.get("Pavement Area (yd²)")
        if pci is None or area is None:
            continue

        segments.append({
            "gisid": d.get("GISID"),
            "street_name": str(name).strip().upper(),
            "from_street": str(d.get("From Street", "")).strip().upper(),
            "to_street": str(d.get("To Street", "")).strip().upper(),
            "length_ft": float(d.get("Pavement Length (ft)", 0) or 0),
            "width_ft": float(d.get("Pavement Width (ft)", 0) or 0),
            "area_sqyd": float(area),
            "pci": float(pci),
            "functional_class": str(d.get("Agency Functional Class", "Local")).strip(),
        })

    wb.close()
    return segments


def compute_segment_costs(segments: list[dict]) -> list[dict]:
    """Compute annual lifecycle cost for each PCI segment."""
    for seg in segments:
        pci = seg["pci"]
        fc = seg["functional_class"]
        cost_table = COST_PER_SQYD[FC_TO_COST_TABLE.get(fc, "urban_local")]

        treatment = "slurry"
        cycle_years = 8
        for lo, hi, treat, cycle in PCI_TREATMENT:
            if lo <= pci <= hi:
                treatment = treat
                cycle_years = cycle
                break

        cost_per_sqyd = cost_table[treatment]
        one_time_cost = seg["area_sqyd"] * cost_per_sqyd
        annual_cost = one_time_cost / cycle_years

        seg["treatment"] = treatment
        seg["cycle_years"] = cycle_years
        seg["cost_per_sqyd"] = cost_per_sqyd
        seg["one_time_cost"] = one_time_cost
        seg["annual_cost"] = annual_cost
        seg["is_local"] = FC_IS_LOCAL.get(fc, True)

    return segments


def normalize_street_name(name: str) -> str:
    """Normalize street name for matching."""
    n = name.upper().strip()
    replacements = {
        " STREET": " ST", " AVENUE": " AVE", " BOULEVARD": " BLVD",
        " DRIVE": " DR", " ROAD": " RD", " LANE": " LN",
        " CIRCLE": " CIR", " COURT": " CT", " PLACE": " PL",
        " WAY": " WAY", " TERRACE": " TER", " TRAIL": " TRL",
    }
    for full, abbr in replacements.items():
        if n.endswith(full):
            n = n[:-len(full)] + abbr
    n = n.replace("  ", " ")
    return n


def match_pci_to_centerlines(
    segments: list[dict],
    streets_gdf: gpd.GeoDataFrame,
) -> gpd.GeoDataFrame:
    """Match PCI segments to street centerline geometries by name.

    Returns a GeoDataFrame of PCI segments with geometries from matching centerlines.
    """
    streets_gdf = streets_gdf.copy()
    streets_gdf["norm_name"] = streets_gdf["RD30FULL"].fillna("").apply(
        lambda x: normalize_street_name(str(x))
    )

    street_index = defaultdict(list)
    for idx, row in streets_gdf.iterrows():
        name = row["norm_name"]
        if name:
            street_index[name].append(idx)

    matched = []
    unmatched = []
    for seg in segments:
        pci_name = normalize_street_name(seg["street_name"])
        candidates = street_index.get(pci_name, [])

        if not candidates:
            words = pci_name.split()
            if len(words) >= 2:
                for stored_name, idxs in street_index.items():
                    if words[0] in stored_name and words[-1] in stored_name:
                        candidates = idxs
                        break

        if candidates:
            best_idx = candidates[0]
            if len(candidates) > 1 and seg["length_ft"] > 0:
                target_len = seg["length_ft"]
                best_diff = float("inf")
                for idx in candidates:
                    cl_len = streets_gdf.loc[idx, "LENGTH"]
                    if cl_len and abs(cl_len - target_len) < best_diff:
                        best_diff = abs(cl_len - target_len)
                        best_idx = idx

            geom = streets_gdf.loc[best_idx, "geometry"]
            seg["geometry"] = geom
            seg["centerline_idx"] = best_idx
            seg["matched_road"] = streets_gdf.loc[best_idx, "RD30FULL"]
            matched.append(seg)
        else:
            unmatched.append(seg)

    print(f"  Matched: {len(matched):,} / {len(segments):,} PCI segments ({100*len(matched)/len(segments):.0f}%)")
    print(f"  Unmatched: {len(unmatched):,}")

    if unmatched:
        unmatched_names = set(normalize_street_name(s["street_name"]) for s in unmatched)
        print(f"  Sample unmatched: {list(unmatched_names)[:10]}")

    gdf = gpd.GeoDataFrame(matched, geometry="geometry", crs=streets_gdf.crs)
    return gdf


def attribute_local_by_frontage(
    local_segments: gpd.GeoDataFrame,
    parcels: gpd.GeoDataFrame,
) -> dict[str, float]:
    """Attribute local street costs to parcels by frontage.

    Buffer each centerline segment, find intersecting parcels,
    measure intersection length as proxy for frontage.
    """
    local_proj = local_segments.to_crs("EPSG:2230")
    parcels_proj = parcels.to_crs("EPSG:2230")

    parcel_sindex = parcels_proj.sindex
    parcel_costs = defaultdict(float)

    for idx, seg in local_proj.iterrows():
        if seg.geometry is None:
            continue

        buffer_dist = 100  # feet — roughly half a residential block depth
        buffered = seg.geometry.buffer(buffer_dist)

        candidate_idxs = list(parcel_sindex.intersection(buffered.bounds))
        if not candidate_idxs:
            continue

        candidates = parcels_proj.iloc[candidate_idxs]
        intersecting = candidates[candidates.intersects(buffered)]

        if len(intersecting) == 0:
            continue

        frontages = []
        for pidx, parcel in intersecting.iterrows():
            boundary = parcel.geometry.boundary
            if boundary is None:
                continue
            intersection = boundary.intersection(buffered)
            frontage = intersection.length if not intersection.is_empty else 0
            frontages.append((pidx, frontage))

        total_frontage = sum(f for _, f in frontages)
        if total_frontage <= 0:
            continue

        annual_cost = seg["annual_cost"]
        for pidx, frontage in frontages:
            share = frontage / total_frontage
            apn = parcels_proj.loc[pidx, "APN"]
            parcel_costs[apn] += annual_cost * share

    return dict(parcel_costs)


def attribute_arterial_by_trips(
    arterial_segments: gpd.GeoDataFrame,
    parcels: gpd.GeoDataFrame,
    revenue_data: dict,
) -> dict[str, float]:
    """Attribute arterial/collector costs by ITE trip generation.

    Each parcel generates trips based on land use type.
    Arterial costs divided proportionally to trip generation.
    """
    total_arterial_cost = float(arterial_segments["annual_cost"].sum())
    print(f"  Total arterial/collector annual cost: ${total_arterial_cost:,.0f}")

    parcel_lookup = {p["apn"]: p for p in revenue_data["parcels"]}

    parcel_trips = {}
    total_trips = 0
    for _, parcel in parcels.iterrows():
        apn = parcel["APN"]
        pdata = parcel_lookup.get(apn, {})
        broad = pdata.get("land_use_broad", "Unknown")
        units = pdata.get("units", 0) or 0
        sqft = pdata.get("living_sqft", 0) or 0

        rate = ITE_TRIPS.get(broad, 1.0)
        if broad in ("SFR", "MFR", "Mixed Use"):
            trips = rate * max(units, 1)
        elif broad in ("Commercial", "Industrial", "Institutional", "Government"):
            trips = rate * max(sqft, 1000) / 1000
        else:
            trips = rate

        parcel_trips[apn] = trips
        total_trips += trips

    parcel_costs = {}
    if total_trips > 0:
        for apn, trips in parcel_trips.items():
            share = trips / total_trips
            parcel_costs[apn] = total_arterial_cost * share

    return parcel_costs


def main():
    print("Loading PCI segments...")
    segments = load_pci_segments()
    print(f"  {len(segments):,} PCI segments loaded")

    print("Computing segment treatment costs...")
    segments = compute_segment_costs(segments)
    total_annual = sum(s["annual_cost"] for s in segments)
    print(f"  Total annual lifecycle cost: ${total_annual:,.0f}")

    local_segs = [s for s in segments if s["is_local"]]
    arterial_segs = [s for s in segments if not s["is_local"]]
    print(f"  Local: {len(local_segs):,} segments (${sum(s['annual_cost'] for s in local_segs):,.0f}/yr)")
    print(f"  Arterial/collector: {len(arterial_segs):,} segments (${sum(s['annual_cost'] for s in arterial_segs):,.0f}/yr)")

    print("\nLoading street centerlines...")
    streets = gpd.read_file(STREETS_FILE)
    print(f"  {len(streets):,} centerline segments")

    print("\nMatching PCI segments to centerlines...")
    matched_local = match_pci_to_centerlines(local_segs, streets)
    matched_arterial = match_pci_to_centerlines(arterial_segs, streets)

    print("\nLoading parcels...")
    parcels = gpd.read_file(PARCELS_FILE)
    print(f"  {len(parcels):,} parcels")

    print("\nLoading revenue data...")
    with open(HERE / "revenue_per_acre.json") as f:
        revenue_data = json.load(f)

    print("\nAttributing local street costs by frontage...")
    print(f"  Processing {len(matched_local):,} matched local segments...")
    local_costs = attribute_local_by_frontage(matched_local, parcels)
    print(f"  Attributed to {len(local_costs):,} parcels")
    print(f"  Total attributed: ${sum(local_costs.values()):,.0f}")

    print("\nAttributing arterial costs by trip generation...")
    arterial_costs = attribute_arterial_by_trips(matched_arterial, parcels, revenue_data)
    print(f"  Attributed to {len(arterial_costs):,} parcels")

    print("\nCombining road costs...")
    all_apns = set(local_costs.keys()) | set(arterial_costs.keys())
    parcel_lookup = {p["apn"]: p for p in revenue_data["parcels"]}

    road_costs = []
    for apn in all_apns:
        local = local_costs.get(apn, 0)
        arterial = arterial_costs.get(apn, 0)
        total_road = local + arterial
        pdata = parcel_lookup.get(apn, {})
        acres = pdata.get("acres", 0)
        revenue = pdata.get("property_tax_city", 0)

        road_costs.append({
            "apn": apn,
            "road_cost_local": round(local, 2),
            "road_cost_arterial": round(arterial, 2),
            "road_cost_total": round(total_road, 2),
            "road_cost_per_acre": round(total_road / acres, 2) if acres > 0.001 else 0,
            "property_tax_city": round(revenue, 2),
            "net_after_roads": round(revenue - total_road, 2),
            "net_per_acre": round((revenue - total_road) / acres, 2) if acres > 0.001 else 0,
            "land_use_broad": pdata.get("land_use_broad", "Unknown"),
            "acres": round(acres, 4),
        })

    road_costs.sort(key=lambda x: -x["road_cost_per_acre"])

    # Summary by land use
    by_category = defaultdict(lambda: {"count": 0, "total_road": 0, "total_tax": 0,
                                        "total_acres": 0, "costs": []})
    for rc in road_costs:
        cat = rc["land_use_broad"]
        by_category[cat]["count"] += 1
        by_category[cat]["total_road"] += rc["road_cost_total"]
        by_category[cat]["total_tax"] += rc["property_tax_city"]
        by_category[cat]["total_acres"] += rc["acres"]
        by_category[cat]["costs"].append(rc["road_cost_per_acre"])

    summary_by_cat = []
    for cat, info in by_category.items():
        costs_arr = np.array(info["costs"])
        summary_by_cat.append({
            "category": cat,
            "count": info["count"],
            "total_acres": round(info["total_acres"], 1),
            "total_road_cost": round(info["total_road"]),
            "total_tax_revenue": round(info["total_tax"]),
            "net_fiscal": round(info["total_tax"] - info["total_road"]),
            "avg_road_cost_per_acre": round(float(costs_arr.mean())) if len(costs_arr) else 0,
            "median_road_cost_per_acre": round(float(np.median(costs_arr))) if len(costs_arr) else 0,
        })
    summary_by_cat.sort(key=lambda x: -x["median_road_cost_per_acre"])

    summary = {
        "methodology": {
            "local_streets": "frontage-based attribution (100ft buffer, proportional to boundary intersection length)",
            "arterials": "ITE trip generation rates by land use type",
            "cost_source": "SCS 2022 treatment costs by PCI bracket",
            "note": "Costs represent lifecycle maintenance need, not current spending",
        },
        "totals": {
            "total_annual_road_cost": round(total_annual),
            "attributed_local": round(sum(local_costs.values())),
            "attributed_arterial": round(sum(arterial_costs.values())),
            "total_attributed": round(sum(local_costs.values()) + sum(arterial_costs.values())),
            "parcels_with_road_cost": len(road_costs),
            "pci_segments_matched_local": len(matched_local),
            "pci_segments_matched_arterial": len(matched_arterial),
        },
        "by_land_use": summary_by_cat,
    }

    output = {"summary": summary, "parcels": road_costs}
    OUTPUT.write_text(json.dumps(output, indent=2))
    SUMMARY_OUTPUT.write_text(json.dumps(summary, indent=2))

    size_mb = OUTPUT.stat().st_size / 1_048_576
    print(f"\n→ {OUTPUT} ({size_mb:.1f} MB)")
    print(f"→ {SUMMARY_OUTPUT}")

    print(f"\n{'=' * 70}")
    t = summary["totals"]
    print(f"Total annual road cost (lifecycle): ${t['total_annual_road_cost']:,}")
    print(f"Attributed local: ${t['attributed_local']:,}")
    print(f"Attributed arterial: ${t['attributed_arterial']:,}")
    print(f"Total attributed: ${t['total_attributed']:,}")
    print(f"Parcels with road cost: {t['parcels_with_road_cost']:,}")

    print(f"\n{'Category':20s} {'Count':>7s} {'Road $/ac':>12s} {'Net $/ac':>12s}")
    print("-" * 55)
    for c in summary_by_cat:
        net_per_ac = round((c["total_tax_revenue"] - c["total_road_cost"]) / c["total_acres"]) if c["total_acres"] > 0 else 0
        print(f"{c['category']:20s} {c['count']:>7,} ${c['median_road_cost_per_acre']:>10,} ${net_per_ac:>10,}")


if __name__ == "__main__":
    main()
